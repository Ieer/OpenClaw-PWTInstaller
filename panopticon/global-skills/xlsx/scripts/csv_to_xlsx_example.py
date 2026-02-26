#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import html
import json
import zipfile
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET


X_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
R_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
NS = {"x": X_NS, "r": R_NS}


def read_csv_rows(input_csv: Path) -> list[list[str]]:
    with input_csv.open("r", encoding="utf-8", newline="") as f:
        reader = csv.reader(f)
        rows = [list(row) for row in reader]
    return rows


def col_to_letters(index_1_based: int) -> str:
    letters = ""
    n = index_1_based
    while n > 0:
        n, rem = divmod(n - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


def is_number(text: str) -> bool:
    try:
        float(text)
        return True
    except Exception:
        return False


def convert_with_openpyxl(rows: list[list[str]], output_xlsx: Path) -> tuple[bool, str]:
    try:
        from openpyxl import Workbook  # type: ignore
        from openpyxl.styles import Font  # type: ignore
    except Exception:
        return False, "openpyxl not installed"

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    for r_idx, row in enumerate(rows, start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            if value is None:
                cell.value = ""
            elif is_number(value) and r_idx > 1:
                number_value = float(value)
                cell.value = int(number_value) if number_value.is_integer() else number_value
            else:
                cell.value = value

    if rows and rows[0]:
        for c_idx in range(1, len(rows[0]) + 1):
            ws.cell(row=1, column=c_idx).font = Font(bold=True)

    ws.freeze_panes = "A2"
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)
    return True, "converted by openpyxl"


def build_styles_xml() -> str:
    return (
        "<?xml version=\"1.0\" encoding=\"UTF-8\" standalone=\"yes\"?>"
        "<styleSheet xmlns=\"http://schemas.openxmlformats.org/spreadsheetml/2006/main\">"
        "<fonts count=\"2\">"
        "<font><sz val=\"11\"/><name val=\"Calibri\"/></font>"
        "<font><b/><sz val=\"11\"/><name val=\"Calibri\"/></font>"
        "</fonts>"
        "<fills count=\"2\">"
        "<fill><patternFill patternType=\"none\"/></fill>"
        "<fill><patternFill patternType=\"gray125\"/></fill>"
        "</fills>"
        "<borders count=\"1\"><border><left/><right/><top/><bottom/><diagonal/></border></borders>"
        "<cellStyleXfs count=\"1\"><xf numFmtId=\"0\" fontId=\"0\" fillId=\"0\" borderId=\"0\"/></cellStyleXfs>"
        "<cellXfs count=\"2\">"
        "<xf numFmtId=\"0\" fontId=\"0\" fillId=\"0\" borderId=\"0\" xfId=\"0\"/>"
        "<xf numFmtId=\"0\" fontId=\"1\" fillId=\"0\" borderId=\"0\" xfId=\"0\" applyFont=\"1\"/>"
        "</cellXfs>"
        "<cellStyles count=\"1\"><cellStyle name=\"Normal\" xfId=\"0\" builtinId=\"0\"/></cellStyles>"
        "</styleSheet>"
    )


def build_content_types_xml() -> str:
    return (
        "<?xml version=\"1.0\" encoding=\"UTF-8\" standalone=\"yes\"?>"
        "<Types xmlns=\"http://schemas.openxmlformats.org/package/2006/content-types\">"
        "<Default Extension=\"rels\" ContentType=\"application/vnd.openxmlformats-package.relationships+xml\"/>"
        "<Default Extension=\"xml\" ContentType=\"application/xml\"/>"
        "<Override PartName=\"/xl/workbook.xml\" ContentType=\"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml\"/>"
        "<Override PartName=\"/xl/worksheets/sheet1.xml\" ContentType=\"application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml\"/>"
        "<Override PartName=\"/xl/styles.xml\" ContentType=\"application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml\"/>"
        "<Override PartName=\"/xl/sharedStrings.xml\" ContentType=\"application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml\"/>"
        "<Override PartName=\"/docProps/core.xml\" ContentType=\"application/vnd.openxmlformats-package.core-properties+xml\"/>"
        "<Override PartName=\"/docProps/app.xml\" ContentType=\"application/vnd.openxmlformats-officedocument.extended-properties+xml\"/>"
        "</Types>"
    )


def build_root_rels_xml() -> str:
    return (
        "<?xml version=\"1.0\" encoding=\"UTF-8\" standalone=\"yes\"?>"
        "<Relationships xmlns=\"http://schemas.openxmlformats.org/package/2006/relationships\">"
        "<Relationship Id=\"rId1\" Type=\"http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument\" Target=\"xl/workbook.xml\"/>"
        "<Relationship Id=\"rId2\" Type=\"http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties\" Target=\"docProps/core.xml\"/>"
        "<Relationship Id=\"rId3\" Type=\"http://schemas.openxmlformats.org/officeDocument/2006/relationships/extended-properties\" Target=\"docProps/app.xml\"/>"
        "</Relationships>"
    )


def build_workbook_xml() -> str:
    return (
        "<?xml version=\"1.0\" encoding=\"UTF-8\" standalone=\"yes\"?>"
        "<workbook xmlns=\"http://schemas.openxmlformats.org/spreadsheetml/2006/main\" "
        "xmlns:r=\"http://schemas.openxmlformats.org/officeDocument/2006/relationships\">"
        "<sheets><sheet name=\"Data\" sheetId=\"1\" r:id=\"rId1\"/></sheets>"
        "</workbook>"
    )


def build_workbook_rels_xml() -> str:
    return (
        "<?xml version=\"1.0\" encoding=\"UTF-8\" standalone=\"yes\"?>"
        "<Relationships xmlns=\"http://schemas.openxmlformats.org/package/2006/relationships\">"
        "<Relationship Id=\"rId1\" Type=\"http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet\" Target=\"worksheets/sheet1.xml\"/>"
        "<Relationship Id=\"rId2\" Type=\"http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles\" Target=\"styles.xml\"/>"
        "<Relationship Id=\"rId3\" Type=\"http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings\" Target=\"sharedStrings.xml\"/>"
        "</Relationships>"
    )


def build_core_xml() -> str:
    return (
        "<?xml version=\"1.0\" encoding=\"UTF-8\" standalone=\"yes\"?>"
        "<cp:coreProperties xmlns:cp=\"http://schemas.openxmlformats.org/package/2006/metadata/core-properties\" "
        "xmlns:dc=\"http://purl.org/dc/elements/1.1/\" "
        "xmlns:dcterms=\"http://purl.org/dc/terms/\" "
        "xmlns:dcmitype=\"http://purl.org/dc/dcmitype/\" "
        "xmlns:xsi=\"http://www.w3.org/2001/XMLSchema-instance\">"
        "<dc:title>CSV to XLSX Example</dc:title>"
        "<dc:creator>OpenClaw XLSX Skill</dc:creator>"
        "</cp:coreProperties>"
    )


def build_app_xml() -> str:
    return (
        "<?xml version=\"1.0\" encoding=\"UTF-8\" standalone=\"yes\"?>"
        "<Properties xmlns=\"http://schemas.openxmlformats.org/officeDocument/2006/extended-properties\" "
        "xmlns:vt=\"http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes\">"
        "<Application>OpenClaw XLSX Skill</Application>"
        "</Properties>"
    )


def build_shared_strings(rows: list[list[str]]) -> tuple[str, dict[str, int]]:
    unique: list[str] = []
    index_map: dict[str, int] = {}

    for r_idx, row in enumerate(rows, start=1):
        for value in row:
            if value is None:
                value = ""
            if is_number(value) and r_idx > 1:
                continue
            if value not in index_map:
                index_map[value] = len(unique)
                unique.append(value)

    si_xml = "".join(f"<si><t>{html.escape(value)}</t></si>" for value in unique)
    sst_xml = (
        "<?xml version=\"1.0\" encoding=\"UTF-8\" standalone=\"yes\"?>"
        "<sst xmlns=\"http://schemas.openxmlformats.org/spreadsheetml/2006/main\" "
        f"count=\"{len(unique)}\" uniqueCount=\"{len(unique)}\">"
        f"{si_xml}</sst>"
    )
    return sst_xml, index_map


def build_sheet_xml(rows: list[list[str]], sst_index: dict[str, int]) -> str:
    if not rows:
        rows = [["empty"]]

    max_col = max(len(r) for r in rows) if rows else 1
    max_row = len(rows)
    dim = f"A1:{col_to_letters(max_col)}{max_row}"

    row_xml_parts: list[str] = []
    for r_idx, row in enumerate(rows, start=1):
        cell_parts: list[str] = []
        for c_idx, value in enumerate(row, start=1):
            ref = f"{col_to_letters(c_idx)}{r_idx}"

            if r_idx > 1 and value is not None and is_number(value):
                cell_parts.append(f"<c r=\"{ref}\"><v>{value}</v></c>")
            else:
                text_value = value if value is not None else ""
                sst_i = sst_index[text_value]
                style_idx = "1" if r_idx == 1 else "0"
                cell_parts.append(f"<c r=\"{ref}\" t=\"s\" s=\"{style_idx}\"><v>{sst_i}</v></c>")

        row_xml_parts.append(f"<row r=\"{r_idx}\">{''.join(cell_parts)}</row>")

    return (
        "<?xml version=\"1.0\" encoding=\"UTF-8\" standalone=\"yes\"?>"
        "<worksheet xmlns=\"http://schemas.openxmlformats.org/spreadsheetml/2006/main\">"
        f"<dimension ref=\"{dim}\"/>"
        "<sheetViews><sheetView workbookViewId=\"0\"><pane ySplit=\"1\" topLeftCell=\"A2\" state=\"frozen\"/></sheetView></sheetViews>"
        f"<sheetData>{''.join(row_xml_parts)}</sheetData>"
        "</worksheet>"
    )


def convert_with_builtin_ooxml(rows: list[list[str]], output_xlsx: Path) -> tuple[bool, str]:
    try:
        output_xlsx.parent.mkdir(parents=True, exist_ok=True)
        sst_xml, sst_index = build_shared_strings(rows)
        sheet_xml = build_sheet_xml(rows, sst_index)

        with zipfile.ZipFile(output_xlsx, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            zf.writestr("[Content_Types].xml", build_content_types_xml())
            zf.writestr("_rels/.rels", build_root_rels_xml())
            zf.writestr("docProps/core.xml", build_core_xml())
            zf.writestr("docProps/app.xml", build_app_xml())
            zf.writestr("xl/workbook.xml", build_workbook_xml())
            zf.writestr("xl/_rels/workbook.xml.rels", build_workbook_rels_xml())
            zf.writestr("xl/styles.xml", build_styles_xml())
            zf.writestr("xl/sharedStrings.xml", sst_xml)
            zf.writestr("xl/worksheets/sheet1.xml", sheet_xml)

        return True, "converted by built-in OOXML fallback"
    except Exception as exc:
        return False, f"built-in fallback failed: {exc}"


def style_check(xlsx_path: Path) -> dict[str, Any]:
    report: dict[str, Any] = {
        "xlsx": str(xlsx_path),
        "ok": False,
        "checks": {},
        "warnings": [],
    }

    if not xlsx_path.exists():
        report["warnings"].append("xlsx file does not exist")
        return report

    try:
        with zipfile.ZipFile(xlsx_path, "r") as zf:
            names = set(zf.namelist())
            report["checks"]["has_workbook_xml"] = "xl/workbook.xml" in names
            report["checks"]["has_sheet1_xml"] = "xl/worksheets/sheet1.xml" in names
            report["checks"]["has_styles_xml"] = "xl/styles.xml" in names
            report["checks"]["has_workbook_rels"] = "xl/_rels/workbook.xml.rels" in names

            if not all(report["checks"].values()):
                report["warnings"].append("missing required workbook parts")
                return report

            workbook_root = ET.fromstring(zf.read("xl/workbook.xml"))
            sheets = workbook_root.findall(".//x:sheets/x:sheet", NS)
            report["checks"]["has_at_least_one_sheet"] = len(sheets) >= 1

            sheet_root = ET.fromstring(zf.read("xl/worksheets/sheet1.xml"))
            rows = sheet_root.findall(".//x:sheetData/x:row", NS)
            report["checks"]["has_rows"] = len(rows) >= 2

            header_cells = sheet_root.findall(".//x:sheetData/x:row[@r='1']/x:c", NS)
            header_style_ok = any(c.attrib.get("s") == "1" for c in header_cells)
            report["checks"]["header_style_applied"] = header_style_ok

            if not header_style_ok:
                report["warnings"].append("header style index=1 not detected on row 1")

            report["ok"] = all(report["checks"].values())
            return report
    except Exception as exc:
        report["warnings"].append(f"style check failed: {exc}")
        return report


def main() -> int:
    parser = argparse.ArgumentParser(description="Minimal CSV -> XLSX example with structure/style checks")
    parser.add_argument("--input", required=True, help="input csv file path")
    parser.add_argument("--output", required=True, help="output xlsx file path")
    parser.add_argument("--report", default="", help="optional json report path")
    args = parser.parse_args()

    input_csv = Path(args.input)
    output_xlsx = Path(args.output)

    if not input_csv.exists():
        print(f"[error] input csv not found: {input_csv}")
        return 1

    rows = read_csv_rows(input_csv)
    if not rows:
        print("[error] input csv is empty")
        return 1

    ok, message = convert_with_openpyxl(rows, output_xlsx)
    if not ok:
        ok, fallback_message = convert_with_builtin_ooxml(rows, output_xlsx)
        message = f"{message}; fallback: {fallback_message}"

    if not ok:
        print("[error] conversion failed")
        print(f"[hint] {message}")
        return 2

    check_report = style_check(output_xlsx)
    check_report["conversion_message"] = message

    print("[ok] generated:", output_xlsx)
    print("[check]", json.dumps(check_report, ensure_ascii=False, indent=2))

    if args.report:
        report_path = Path(args.report)
        report_path.parent.mkdir(parents=True, exist_ok=True)
        report_path.write_text(json.dumps(check_report, ensure_ascii=False, indent=2), encoding="utf-8")
        print("[ok] report:", report_path)

    return 0 if check_report.get("ok") else 3


if __name__ == "__main__":
    raise SystemExit(main())
