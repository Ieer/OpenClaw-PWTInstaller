#!/usr/bin/env python3
"""生成 WCQ 下周工作计划 Excel — 专业版"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle
)
from openpyxl.utils import get_column_letter
from datetime import datetime

wb = Workbook()

# ── 颜色定义 ──
HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")   # 深蓝
P0_FILL       = PatternFill("solid", fgColor="FCE4EC")   # 浅红
P1_FILL       = PatternFill("solid", fgColor="FFF3E0")   # 浅橙
BELLA_FILL    = PatternFill("solid", fgColor="E8F5E9")   # 浅绿
TRAIN_FILL    = PatternFill("solid", fgColor="E3F2FD")   # 浅蓝
REVIEW_FILL   = PatternFill("solid", fgColor="F3E5F5")   # 浅紫
SUMMARY_FILL  = PatternFill("solid", fgColor="FFF8E1")   # 浅黄
ODD_ROW       = PatternFill("solid", fgColor="F5F7FA")   # 极浅灰-条纹
RED_FONT      = Font(color="D32F2F", bold=True)
WHITE_FONT    = Font(color="FFFFFF", bold=True, size=10)
NORMAL_FONT   = Font(name="微软雅黑", size=10)
BOLD_FONT     = Font(name="微软雅黑", size=10, bold=True)
TITLE_FONT    = Font(name="微软雅黑", size=14, bold=True, color="1F4E79")
SUBTITLE_FONT = Font(name="微软雅黑", size=10, color="666666")
THIN_BORDER   = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

# ═══════════════════════════════
#  Sheet 1: 每日时间表
# ═══════════════════════════════
ws = wb.active
ws.title = "下周日程"

# 列宽
col_widths = [5, 12, 50, 30, 12, 14]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ── 标题行 ──
ws.merge_cells("A1:F1")
ws["A1"] = "🗓  WCQ 数位工具 — 下周工作计划 (2026.05.18 ~ 05.22)"
ws["A1"].font = TITLE_FONT
ws["A1"].alignment = Alignment(vertical="center")
ws.row_dimensions[1].height = 36

ws.merge_cells("A2:F2")
ws["A2"] = "PMO: Ieer Qin  |  协助: Bella Wu  |  生成时间: 2026-05-16"
ws["A2"].font = SUBTITLE_FONT
ws.row_dimensions[2].height = 22

# ── 表头 ──
headers = ["#", "时间", "任务", "备注", "优先级", "标签"]
for i, h in enumerate(headers, 1):
    cell = ws.cell(row=3, column=i, value=h)
    cell.font = WHITE_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER
ws.row_dimensions[3].height = 28

# ── 数据 ──
data = [
    # (day_label, time, task, note, priority, tag)
    # === 周一 5/18 ===
    ("05/18 周一", "09:00-09:30", "Swap回收", "sudo swapoff -a && swapon -a", "🔴 高", "系统维护"),
    ("05/18 周一", "09:30-10:00", "专业训最后20%校对", "收尾验收，确认可上架", "🔴 高", "开发-P0"),
    ("05/18 周一", "10:00-10:30", "Python培训反馈收集", "联系学员收集第3天反馈", "🟢 中", "培训"),
    ("05/18 周一", "13:30-14:00", "🔄 Bella交接 #1 — 项目Delay追踪", "教跑COE系统→红黄灯清单", "🟢 中", "Bella交接"),
    ("05/18 周一", "14:00-14:30", "🔄 Bella交接 #2 — SAP进度追踪", "移交联系人清单，教进度表", "🟢 中", "Bella交接"),
    ("05/18 周一", "14:30-15:00", "🔄 Bella交接 #3 — 评审跟催", "教看状态码，跟催模板", "🟢 中", "Bella交接"),
    ("05/18 周一", "15:00-15:30", "NPI RFQ资料确认", "资料已到→启动开发/未到→确认时间", "🟡 中", "确认"),
    ("05/18 周一", "16:00-17:00", "SAP各部门邮件整理", "已回复/未回复清单，催一次", "🟡 中", "管理"),
    # === 周二 5/19 ===
    ("05/19 周二", "09:00-11:00", "✅ 专业训上架部署", "校对→测试→部署上架", "🔴 高", "开发-P0"),
    ("05/19 周二", "11:00-12:00", "PCBA钢板整合测试启动", "拉通各模块，开始整合", "🟡 中", "开发-P1"),
    ("05/19 周二", "14:00-15:30", "PCBA钢板开发", "整合测试，记录问题清单", "🟡 中", "开发-P1"),
    ("05/19 周二", "16:00-17:00", "📖 Bella培训 — Week3 项目管理", "数位项目管理(需求→平台运作)", "🟢 中", "培训"),
    ("05/19 周二", "17:00-17:30", "交响乐 — 报告迭代", "Columbus报告优化(距5/29剩10天)", "🔴 高", "活动"),
    # === 周三 5/20 ===
    ("05/20 周三", "09:00-10:00", "🔴 交响乐报告Approve", "Deadline: 内部Approve截止", "🔴 高", "活动-截止"),
    ("05/20 周三", "10:00-12:00", "PCBA钢板开发", "整合测试集中攻坚", "🟡 中", "开发-P1"),
    ("05/20 周三", "14:00-15:00", "SAP各部门进度确认", "跟踪周一催过的邮件回复", "🟡 中", "管理"),
    ("05/20 周三", "15:00-16:00", "PCBA钢板开发 #2", "第二轮整合测试", "🟡 中", "开发-P1"),
    ("05/20 周三", "16:00-16:30", "Bella — 本周任务Check", "检查Delay/SAP/跟催执行情况", "🟢 中", "管理"),
    # === 周四 5/21 ===
    ("05/21 周四", "09:00-12:00", "PCBA钢板集中攻坚", "完整半天给整合测试", "🟡 中", "开发-P1"),
    ("05/21 周四", "14:00-15:00", "NPI RFQ Agent开发启动", "框架搭建+数据导入(若资料已到)", "🟢 中", "开发"),
    ("05/21 周四", "15:00-15:30", "📖 Bella培训 — Week3 案例", "数位项目管理案例说明", "🟢 中", "培训"),
    ("05/21 周四", "15:30-17:00", "Bella实操 — SAP进度独立跑", "Bella独立跑一次，Ieer审核", "🟢 中", "Bella实操"),
    ("05/21 周四", "17:00-17:30", "交响乐 — 报告定稿确认", "确认提交准备(5/25截止)", "🔴 高", "活动"),
    # === 周五 5/22 ===
    ("05/22 周五", "09:00-11:00", "PCBA钢板开发", "整合测试收尾，评估本周进度", "🟡 中", "开发-P1"),
    ("05/22 周五", "11:00-12:00", "交响乐 — 报告提交准备", "内容终版确认，准备提交", "🔴 高", "活动"),
    ("05/22 周五", "14:00-14:30", "Bella — 本周工作Review", "检查清单/进度表/跟催，反馈调整", "🟢 中", "管理"),
    ("05/22 周五", "14:30-15:00", "Bella — 下周培训预告", "Week4课程提醒(数位项目挖掘)", "🟢 低", "培训"),
    ("05/22 周五", "15:00-15:30", "专业训上线确认", "部署完成确认，记录备案", "🟢 低", "验收"),
    ("05/22 周五", "15:30-16:00", "本周工作盘点", "对照计划逐项Check", "🟢 低", "总结"),
    ("05/22 周五", "16:00-16:30", "下周优先级预览", "NPI RFQ / PCBA / 6/1课程通知", "🟢 低", "总结"),
    ("05/22 周五", "16:30-17:00", "🔴 CW报告实战演练", "交响乐Columbus报告彩排", "🔴 高", "活动-截止"),
]

tag_fill = {
    "开发-P0": P0_FILL,
    "开发-P1": P1_FILL,
    "Bella交接": BELLA_FILL,
    "Bella实操": BELLA_FILL,
    "培训": TRAIN_FILL,
    "活动-截止": P0_FILL,
    "活动": P1_FILL,
    "管理": REVIEW_FILL,
    "确认": SUMMARY_FILL,
    "总结": SUMMARY_FILL,
    "验收": SUMMARY_FILL,
    "系统维护": ODD_ROW,
}

row = 4
for i, (day, time, task, note, priority, tag) in enumerate(data, 1):
    ws.cell(row=row, column=1, value=i).font = NORMAL_FONT
    ws.cell(row=row, column=1).alignment = CENTER

    ws.cell(row=row, column=2, value=day).font = BOLD_FONT
    ws.cell(row=row, column=2).alignment = Alignment(vertical="top")

    ws.cell(row=row, column=3, value=time).font = NORMAL_FONT
    ws.cell(row=row, column=3).alignment = Alignment(vertical="top")

    ws.cell(row=row, column=4, value=task).font = BOLD_FONT
    ws.cell(row=row, column=4).alignment = WRAP

    ws.cell(row=row, column=5, value=note).font = Font(name="微软雅黑", size=9, color="666666")
    ws.cell(row=row, column=5).alignment = WRAP

    pcell = ws.cell(row=row, column=6, value=priority)
    pcell.font = RED_FONT if "🔴" in priority else Font(name="微软雅黑", size=10, bold=True)
    pcell.alignment = CENTER

    tcell = ws.cell(row=row, column=7, value=tag)
    tcell.font = Font(name="微软雅黑", size=9, bold=True)
    tcell.alignment = CENTER
    if tag in tag_fill:
        tcell.fill = tag_fill[tag]

    # 边框
    for c in range(1, 8):
        ws.cell(row=row, column=c).border = THIN_BORDER

    ws.row_dimensions[row].height = 32
    row += 1

# Freeze panes
ws.freeze_panes = "A4"

# ═══════════════════════════════
#  Sheet 2: 周目标
# ═══════════════════════════════
ws2 = wb.create_sheet("周目标")

ws2.column_dimensions["A"].width = 5
ws2.column_dimensions["B"].width = 18
ws2.column_dimensions["C"].width = 50
ws2.column_dimensions["D"].width = 18

ws2.merge_cells("A1:D1")
ws2["A1"] = "🎯 周目标 (2026.05.18 ~ 05.22)"
ws2["A1"].font = TITLE_FONT
ws2.row_dimensions[1].height = 36

goal_headers = ["#", "目标", "验收标准", "优先级"]
for i, h in enumerate(goal_headers, 1):
    cell = ws2.cell(row=2, column=i, value=h)
    cell.font = WHITE_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER

goals = [
    ("专业训上架", "COE平台可查，已部署上线", "🔴 P0"),
    ("交响乐报告Approve", "5/20前内部确认通过", "🔴 P0"),
    ("PCBA钢板整合测试50%", "问题清单已记录，修复过半", "🟡 P1"),
    ("Bella三件事交接完成", "Delay追踪+SAP进度+评审跟催可独立执行", "🟡 P1"),
    ("NPI RFQ确认", "明确资料到位时间或启动开发", "🟢 P2"),
    ("Python培训反馈收集", "至少收集3人反馈", "🟢 P2"),
]

for i, (goal, std, pri) in enumerate(goals, 1):
    r = i + 2
    ws2.cell(row=r, column=1, value=i).font = NORMAL_FONT
    ws2.cell(row=r, column=1).alignment = CENTER
    ws2.cell(row=r, column=2, value=goal).font = BOLD_FONT
    ws2.cell(row=r, column=3, value=std).font = NORMAL_FONT
    ws2.cell(row=r, column=3).alignment = WRAP
    pc = ws2.cell(row=r, column=4, value=pri)
    pc.font = RED_FONT if "P0" in pri else Font(name="微软雅黑", size=10, bold=True)
    pc.alignment = CENTER
    for c in range(1, 5):
        ws2.cell(row=r, column=c).border = THIN_BORDER
    ws2.row_dimensions[r].height = 28

# ═══════════════════════════════
#  Sheet 3: Bella 本周培训安排
# ═══════════════════════════════
ws3 = wb.create_sheet("Bella培训安排")

ws3.column_dimensions["A"].width = 5
ws3.column_dimensions["B"].width = 14
ws3.column_dimensions["C"].width = 14
ws3.column_dimensions["D"].width = 45
ws3.column_dimensions["E"].width = 20

ws3.merge_cells("A1:E1")
ws3["A1"] = "📖 Bella 培训安排 — Week3 (5/18~5/22)"
ws3["A1"].font = TITLE_FONT
ws3.row_dimensions[1].height = 36

b_headers = ["#", "日期", "时间", "课程内容", "角色"]
for i, h in enumerate(b_headers, 1):
    cell = ws3.cell(row=2, column=i, value=h)
    cell.font = WHITE_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER

bella_data = [
    ("周一", "全天", "交接：Delay追踪/SAP进度/评审跟催", "带教"),
    ("周二", "16:00-17:00", "数位项目管理（需求→平台运作）", "授课"),
    ("周三", "16:00-16:30", "本周任务Check", "审核"),
    ("周四", "15:30-17:00", "SAP进度追踪独立实操", "实操+审核"),
    ("周四", "16:00-16:30", "数位项目管理 — 案例说明", "授课"),
    ("周五", "14:00-15:00", "本周工作Review + 下周预告", "总结"),
]

for i, (d, t, c, role) in enumerate(bella_data, 1):
    r = i + 2
    ws3.cell(row=r, column=1, value=i).font = NORMAL_FONT
    ws3.cell(row=r, column=1).alignment = CENTER
    ws3.cell(row=r, column=2, value=d).font = BOLD_FONT
    ws3.cell(row=r, column=3, value=t).font = NORMAL_FONT
    ws3.cell(row=r, column=4, value=c).font = NORMAL_FONT
    ws3.cell(row=r, column=4).alignment = WRAP
    rc = ws3.cell(row=r, column=5, value=role)
    rc.font = Font(name="微软雅黑", size=10, bold=True)
    rc.alignment = CENTER
    if role == "授课":
        rc.fill = TRAIN_FILL
    elif role == "实操+审核":
        rc.fill = BELLA_FILL
    elif role == "带教":
        rc.fill = SUMMARY_FILL
    for c2 in range(1, 6):
        ws3.cell(row=r, column=c2).border = THIN_BORDER
    ws3.row_dimensions[r].height = 26

# ── 保存 ──
out_path = "/mnt/usb/WCQ_下周工作计划_2026-05-18.xlsx"
wb.save(out_path)
print(f"✅ 已保存: {out_path}")
